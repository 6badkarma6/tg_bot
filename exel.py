"""Модули для бота"""


# import logger as log
class Modules:
    def __init__(self):
        from bs4 import BeautifulSoup as BeautifulSoup
        self.BeautifulSoup = BeautifulSoup
        import json as json
        self.json = json
        import openpyxl as openpyxl
        self.openpyxl = openpyxl

    @staticmethod
    def lon(data) -> str:
        string = 'Список групп'
        for i in range(len(data)):
            string = string + data[i] + " "
        return string

    def rw(self=None, group: str = "self") -> str:
        """Вывод строки с рассписанием группы"""
        groups = self.ajson("timetable.json")
        return groups[group]

    def time_grs(self):
        list_grs = (self.ajson('bin.json'))['grs']
        time = dict(map(self.strings, list_grs))
        with open('timetable.json', 'w') as file:
            self.json.dump(time, file, indent=2)
        return True

    def strings(self, i):
        data = self.ajson()['sso']
        return i, self.message(self.nn(lists=self.message_dt(gr=i, load_file=data[0])))

    # Фильтр
    @staticmethod
    def nn(lists: list) -> list:
        """Фильтрует от пустых строк и добавляет недостоющие"""
        for j in range(len(lists)):
            for k in range(len(lists[j])):
                if lists[j][k] == 'None' and j != 3:
                    lists[j][k] = lists[j - 1][k]
        for j in range(len(lists)):
            for k in range(len(lists[j])):
                if lists[j][k] == 'None':
                    lists[j][k] = '     '
        return lists

    # Форматирует список в текст
    @staticmethod
    def message(data: list) -> str:
        """Преобразует список в текст"""
        return ''.join([''.join(dt) for dt in data])

    # Формирует список
    def message_dt(self, gr: str, load_file: str, file: str = 'data.json') -> list:
        """Формирует список из .xlsx файла по конфигурации data.json"""
        import openpyxl
        wb = openpyxl.load_workbook(load_file)
        sh = wb.active
        data = self.ajson(file)
        message_data = []
        xd = data[gr]['x']
        yd = data[gr]['y']
        for y in yd:
            dt = []
            for x in xd:
                if str(sh[f'{x}{y}'].value) == 'None' and x != data[gr]['x'][0] and y != data[gr]['y'][0]:
                    dt.append(str(sh[f'{x}{y}'].value))
                elif str(sh[f'{x}{y}'].value) == 'None':
                    dt.append('     ')
                else:
                    dt.append(str(sh[f'{x}{y}'].value) + '    ')
            dt.append('\n')
            message_data.append(dt)
        return message_data

    # Достаёт из исходного кода нужные ссылки и названия
    def pr(self, bol: bool = False):
        """Достаёт из исходного кода нужные ссылки и названия"""

        if bol:
            import os
            os.system("curl -o page_source.html https://vgke.by/raspisanie-zanjatij/")
        with open(file='page_source.html', mode='r', encoding='utf-8') as html:
            print('pr start')
            soup = self.BeautifulSoup(html, 'html.parser')
            http = []
            names = []
            sso = []
            http_sso = []
            for link in soup.find_all('iframe'):
                strings = (link.get('src').split('='))[1]
                http.append(strings)
                name = (strings.split('/'))[-1]
                names.append(name)
                if 'den-sso' in name:
                    http_sso.append(strings)
                    sso.append(name)
                print(http, '\n', http_sso, '\n', names, '\n', sso)
            print('pr finish')
            with open('conf.json', 'w') as write_file:
                self.json.dump({'http': http,
                                'http_sso': http_sso,
                                'names': names,
                                'sso': sso}, write_file, indent=2)
            return True

    # Сохраняет файл
    def save(self, pb=False):
        """Сохраняет файл"""
        import os
        if pb:
            self.pr()
            print('pr True')
        print('start...')
        data = self.ajson()
        os.system("curl -O " + data['http_sso'][0])
        print('...finish')

    # Импортирует данные файла conf.json
    def ajson(self=None, file: str = 'conf.json') -> dict:
        """Импортирует данные"""
        with open(file, 'r', encoding="utf-8") as file:
            data = self.json.load(file)
        return data

    # Сортирует exel файл
    def search(self, file: str, gr: str) -> tuple[int, int]:
        """Сортирует exel файл"""
        workbook = self.openpyxl.load_workbook(file)
        open_file = workbook.active
        for gr_x in range(1, open_file.max_row):
            for gr_y in range(1, open_file.max_column):
                if open_file.cell(row=gr_x, column=gr_y).value == gr:
                    return gr_x, gr_y

    @staticmethod
    def g(x: int, y: int):
        return x + y

    # Состовляет столбцы и строки

    def ls(self, gr_x: int, gr_y1: int, gr_y2, x: int, y: int, y0: int = 3, k: int = 1) -> tuple[list[str], list[str]]:
        # [gr_x, 3, x, y], [gr_x, gr_y1, x, gr_y2]
        with open('bin.json', 'r') as file:
            data = self.json.load(file)
        abc = data['abc']
        row = []
        column = []
        r_glob = x - gr_x + k
        c_timing = y - y0 + k
        c_gr = gr_y2 - gr_y1 + k
        for n in range(0, r_glob):
            row.append(str(self.g(gr_x, n)))
        for i in range(0, c_timing):
            column.append(abc[self.g(y0, i)])
        for j in range(0, c_gr):
            column.append(abc[self.g(gr_y1, j)])
        return column, row

    def filters(self, file: str, gr_x: int, gr_y: int):
        workbook = self.openpyxl.load_workbook(file)
        open_file = workbook.active
        x = gr_x
        y = 3
        for i in range(14):
            x += 1
            if str(open_file.cell(row=x, column=y).value) == '' or str(open_file.cell(row=x, column=y).value) == ' ':
                x -= 1
        for i in range(4):
            if str(open_file.cell(row=gr_x, column=y).value) == 'None':
                break
            y += 1
        if str(open_file.cell(row=gr_x, column=self.g(1, gr_y)).value) == 'None' \
                and str(open_file.cell(row=gr_x, column=self.g(3, gr_y)).value) == 'None':
            gr_y += 3
        elif str(open_file.cell(row=gr_x, column=self.g(1, gr_y)).value) == 'None' \
                and str(open_file.cell(row=gr_x, column=self.g(3, gr_y)).value) != 'None' or \
                str(open_file.cell(row=gr_x, column=self.g(1, gr_y)).value) != 'None' \
                and str(open_file.cell(row=gr_x, column=self.g(2, gr_y)).value) == 'None':
            gr_y += 2
        else:
            gr_y += 1
        return x, y, gr_y

    @staticmethod
    def info(message, name):
        import datetime
        return str(datetime.datetime.fromtimestamp(message.date)) + ": " \
               + str(name.from_user.first_name) + " " \
               + str(name.from_user.id)

    # Формирует конфигурационный файл data.json
    def data_conf(self):
        conf = self.ajson()
        binarn = self.ajson('bin.json')
        grs = binarn['grs']
        file = conf['sso'][0]
        with open(file='data.json', mode='w', encoding='utf-8') as dump_file:
            dump = {}
            for gr in grs:
                print("start ", gr)
                gr_x, gr_y1 = self.search(file, gr)
                x, y, gr_y2 = self.filters(file, gr_x, gr_y1)
                column, row = self.ls(gr_x=gr_x, gr_y1=gr_y1, x=x, y=y, gr_y2=gr_y2)
                dt = dict(x=column, y=row)
                print(dt, end='\n\n')
                dump[gr] = dt
            print(dump)
            self.json.dump(obj=dump, fp=dump_file, indent=4)
        return self.time_grs()

    # Считывет токин бота
    @staticmethod
    def data(i: bool = True) -> str:
        if i:
            with open('tokens/bot', 'r') as data_bot:
                return data_bot.read(46)
        else:
            with open('tokens/badwr', 'r') as data_bot:
                return data_bot.read(46)


class Users:
    def __init__(self):
        self.db = None

    def read_user(self):
        db = {}
        import shelve as sh
        with sh.open("user/user_id", flag='r') as file:
            for i in file:
                db[i] = file[i]
        self.db = db

    def check_id(self, user_id: str) -> bool:
        if user_id in self.db:
            return True
        else:
            return False

    @staticmethod
    def add_user(user_id: str, first_name: str, group: str):
        import shelve as sh
        with sh.open("user/user_id", flag='c') as file:
            file[user_id] = [first_name, group]

    @staticmethod
    def del_user(user_id: str):
        import shelve as sh
        with sh.open("user/user_id", flag='c') as file:
            del file[user_id]

    def post_user(self):
        return self.db

# if __name__ == "__main__":
# import json
# file = "den-sso-1.xlsx"
# gr = str(input('>>'))
# gr_x, gr_y1 = search(file, gr)
# x, y, gr_y2 = filter(file, gr_x, gr_y1, gr)
# print([gr_x, 3, x, y], [gr_x, gr_y1, x, gr_y2])
# column, row = ls(gr_x=gr_x, gr_y1=gr_y1, x=x, y=y, gr_y2=gr_y2)
# print(column, row)
# src()
# pr()
# save()
# data_conf()
# print(rw('ПО-1'))
# print(user(1234))
# inp = int(input('>>'))
# print(user(inp))
